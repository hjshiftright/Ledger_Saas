"""Inspect SBI new-format encrypted files after decryption."""
import sys
import io
import getpass
import os

sys.path.insert(0, "src")

password = getpass.getpass("Enter SBI file password: ")

import msoffcrypto  # noqa: E402
import fitz  # noqa: E402  (PyMuPDF)
import pdfplumber  # noqa: E402

SBI_DIR = "src/samples/sbi"

# ── New format PDFs ───────────────────────────────────────────────────────────
pdf_files = [
    os.path.join(SBI_DIR, "AccountStatement_24032026_094359.pdf"),
    os.path.join(SBI_DIR, "AccountStatement_24032026_094432.pdf"),
]

for path in pdf_files:
    fname = os.path.basename(path)
    with open(path, "rb") as f:
        raw = f.read()
    try:
        doc = fitz.open(stream=raw, filetype="pdf")
        if doc.needs_pass:
            ok = doc.authenticate(password)
            if not ok:
                print(f"[{fname}] Wrong password")
                continue
        print(f"\n=== PDF: {fname}  ({doc.page_count} pages) ===")
        text = doc[0].get_text()
        print(text[:2000])
    except Exception as e:
        print(f"[{fname}] Error: {e}")

# ── New format PDFs via pdfplumber (table extraction) ─────────────────────────
print("\n\n=== pdfplumber table extraction on first new PDF ===")
path = os.path.join(SBI_DIR, "AccountStatement_24032026_094359.pdf")
with open(path, "rb") as f:
    raw = f.read()
try:
    # Decrypt first using PyMuPDF → write decrypted bytes → open with pdfplumber
    doc = fitz.open(stream=raw, filetype="pdf")
    if doc.needs_pass:
        doc.authenticate(password)
    dec_buf = io.BytesIO(doc.write())
    with pdfplumber.open(dec_buf) as pdf:
        for page in pdf.pages[:2]:
            print(f"\n--- Page {page.page_number} ---")
            tables = page.extract_tables()
            print(f"Tables found: {len(tables)}")
            for tbl in tables[:1]:
                for row in tbl[:6]:
                    print(row)
except Exception as e:
    print(f"pdfplumber error: {e}")

# ── New format XLSX ───────────────────────────────────────────────────────────
xlsx_files = [
    os.path.join(SBI_DIR, "AccountStatement_24032026_094520.xlsx"),
    os.path.join(SBI_DIR, "AccountStatement_24032026_094538.xlsx"),
]

for path in xlsx_files:
    fname = os.path.basename(path)
    with open(path, "rb") as f:
        raw = f.read()
    try:
        office = msoffcrypto.OfficeFile(io.BytesIO(raw))
        office.load_key(password=password)
        dec = io.BytesIO()
        office.decrypt(dec)
        dec.seek(0)
        import openpyxl
        wb = openpyxl.load_workbook(dec, data_only=True)
        ws = wb.active
        print(f"\n=== XLSX: {fname}  (sheet: {ws.title}, max_row: {ws.max_row}) ===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = list(row)
            if any(v is not None for v in vals):
                print(f"  Row {i:3d}: {vals}")
            if i >= 35:
                print("  ...")
                break
    except Exception as e:
        print(f"[{fname}] Error: {e}")

print("\nDone.")
