"""Comprehensive tests for SBI parser — all format combinations.

Tests cover:
  - SbiPdfParser: table-based extraction from pdfplumber table fixtures
  - SbiPdfParser: text-layer parsing (parse_text_content)
  - GenericCsvParser: old SBI pseudo-XLS (tab-separated text with 'quoted' cells)
  - GenericCsvParser: SBI XLSX (regular XLSX with preamble rows)
  - Detector: old SBI XLS header scan, new-format filename patterns
  - _rows_from_tables: cell normalisation (embedded newlines, empty amounts, ref col)

All tests use in-memory fixtures — no disk I/O (PDF/XLS files not required).
"""

from __future__ import annotations

import io
import uuid

import pytest

from core.models.enums import ExtractionMethod, SourceType, TxnTypeHint, FileFormat
from modules.parser.parsers.sbi_pdf import SbiPdfParser, _normalise_cell, _infer_txn_type
from modules.parser.parsers.generic_csv import GenericCsvParser
from modules.parser.detector import SourceDetector


# ── Shared helpers ────────────────────────────────────────────────────────────

@pytest.fixture()
def parser() -> SbiPdfParser:
    return SbiPdfParser()


@pytest.fixture()
def csv_parser() -> GenericCsvParser:
    return GenericCsvParser()


@pytest.fixture()
def batch_id() -> str:
    return str(uuid.uuid4())


@pytest.fixture()
def detector() -> SourceDetector:
    return SourceDetector()


# ── Table extraction fixtures ─────────────────────────────────────────────────
# Mirrors pdfplumber output: list[list[list[str]]] → tables → rows

def _make_table(data_rows: list[list[str | None]]) -> list[list[str | None]]:
    """Prepend standard SBI header to data rows."""
    header = ["Txn Date", "Value\nDate", "Description", "Ref No./Cheque\nNo.", "Debit", "Credit", "Balance"]
    return [header] + data_rows


_TABLE_SINGLE_CREDIT = _make_table([
    ["1 Apr 2025", "1 Apr 2025", "BY TRANSFER-\nNEFT*ICIC0SF0002*ICIN10917\n6532615*DANTURTI", "TRANSFER\nFROM\n99509044300", "", "5,000.00", "27,527.17"],
])

_TABLE_SINGLE_DEBIT = _make_table([
    ["3 Apr 2025", "3 Apr 2025", "TO TRANSFER-\nUPI/DR/509399466100/MACHA\nRLA/UTIB/sphomeocli/UPI-", "TRANSFER TO\n4897694162092", "5,800.00", "", "21,729.17"],
])

_TABLE_MULTIPLE = _make_table([
    ["1 Apr 2025", "1 Apr 2025", "BY TRANSFER-\nNEFT*ICIC0SF0002\nDANTURTI", "TRANSFER\nFROM\n99509044300", "", "5,000.00", "27,527.17"],
    ["3 Apr 2025", "3 Apr 2025", "TO TRANSFER-\nUPI/DR/509399466100/MACHA", "TRANSFER TO\n4897694162092", "5,800.00", "", "21,729.17"],
    ["9 Apr 2025", "9 Apr 2025", "TO TRANSFER-\nIMPS/100327963734/NETFLI", "TRANSFER TO\n4897693162093", "199.00", "", "21,530.17"],
    ["9 Apr 2025", "9 Apr 2025", "CASH WD ATM-SBI/001 BANGALORE", "ATM00001234", "2,000.00", "", "19,530.17"],
])

_TABLE_SPLIT_DATE = _make_table([
    # Dates with embedded newlines as produced by pdfplumber for split cells
    ["13 May\n2025", "13 May\n2025", "TO TRANSFER-UPI/DR/540933261335/SELVA", "TRANSFER TO\n4897692162094", "160.00", "", "25,178.17"],
])

_TABLE_EMPTY_AMOUNTS = _make_table([
    # A row with both debit and credit empty — should get row_confidence 0.5
    ["15 Jun 2025", "15 Jun 2025", "MEMO ENTRY / ADJUSTMENT", "", None, None, "25,000.00"],
])

# Two-page scenario: two table objects (pdfplumber gives one table per page)
_TABLE_PAGE2 = _make_table([
    ["1 May 2025", "1 May 2025", "BY TRANSFER-NEFT*DANTURTI", "TRANSFER\nFROM", "", "5,000.00", "26,237.17"],
])

_TABLES_TWO_PAGES = [_TABLE_MULTIPLE, _TABLE_PAGE2]


# ═══════════════════════════════════════════════════════════════════════════════
# 1. SbiPdfParser — TABLE_EXTRACTION via _rows_from_tables
# ═══════════════════════════════════════════════════════════════════════════════

class TestTableExtraction:
    def test_single_credit_row(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_SINGLE_CREDIT])
        assert len(rows) == 1
        r = rows[0]
        assert r.raw_credit == "5,000.00"
        assert r.raw_debit is None
        assert r.raw_balance == "27,527.17"

    def test_single_debit_row(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_SINGLE_DEBIT])
        assert len(rows) == 1
        r = rows[0]
        assert r.raw_debit == "5,800.00"
        assert r.raw_credit is None

    def test_multiple_rows_count(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_MULTIPLE])
        assert len(rows) == 4

    def test_row_numbers_sequential_across_pages(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, _TABLES_TWO_PAGES)
        for i, row in enumerate(rows, start=1):
            assert row.row_number == i

    def test_split_date_normalised(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_SPLIT_DATE])
        assert len(rows) == 1
        assert rows[0].raw_date == "13 May 2025"

    def test_embedded_newlines_stripped_from_narration(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_SINGLE_CREDIT])
        narr = rows[0].raw_narration
        assert "\n" not in narr
        assert "BY TRANSFER" in narr

    def test_ref_number_extracted(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_SINGLE_CREDIT])
        assert rows[0].raw_reference is not None
        assert "99509044300" in rows[0].raw_reference

    def test_source_type_sbi_bank(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_MULTIPLE])
        assert all(r.source_type == SourceType.SBI_BANK for r in rows)

    def test_batch_id_propagated(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_SINGLE_CREDIT])
        assert all(r.batch_id == batch_id for r in rows)

    def test_empty_amounts_get_low_confidence(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_EMPTY_AMOUNTS])
        assert len(rows) == 1
        assert rows[0].row_confidence == 0.5

    def test_valid_amounts_get_high_confidence(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_SINGLE_CREDIT])
        assert rows[0].row_confidence >= 0.9

    def test_parser_version(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_SINGLE_CREDIT])
        assert rows[0].parser_version == "1.2"

    def test_empty_table_returns_empty(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [[]])
        assert rows == []

    def test_table_without_date_col_skipped(self, parser, batch_id):
        bad_table = [["Foo", "Bar"], ["val1", "val2"]]
        rows = SbiPdfParser._rows_from_tables(batch_id, [bad_table])
        assert rows == []

    def test_header_row_not_included(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_SINGLE_CREDIT])
        # No row should have "Txn Date" as its raw_date
        assert not any(r.raw_date.lower() == "txn date" for r in rows)

    def test_two_page_total_count(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, _TABLES_TWO_PAGES)
        assert len(rows) == 5  # 4 from page 1 + 1 from page 2


# ═══════════════════════════════════════════════════════════════════════════════
# 2. SbiPdfParser — TABLE_EXTRACTION build_result and confidence
# ═══════════════════════════════════════════════════════════════════════════════

class TestTableExtractionResult:
    def test_extract_method_tagged(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_MULTIPLE])
        result = parser._build_result(batch_id, rows, ExtractionMethod.TABLE_EXTRACTION)
        assert result.method == ExtractionMethod.TABLE_EXTRACTION

    def test_confidence_positive(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_MULTIPLE])
        result = parser._build_result(batch_id, rows, ExtractionMethod.TABLE_EXTRACTION)
        assert result.confidence > 0.0

    def test_zero_rows_gives_low_confidence(self, parser, batch_id):
        result = parser._build_result(batch_id, [], ExtractionMethod.TABLE_EXTRACTION)
        assert result.confidence < 0.3

    def test_metadata_dates_populated(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_MULTIPLE])
        result = parser._build_result(batch_id, rows, ExtractionMethod.TABLE_EXTRACTION)
        assert result.metadata.statement_from is not None
        assert result.metadata.statement_to is not None

    def test_total_rows_in_metadata(self, parser, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_TABLE_MULTIPLE])
        result = parser._build_result(batch_id, rows, ExtractionMethod.TABLE_EXTRACTION)
        assert result.metadata.total_rows_found == 4


# ═══════════════════════════════════════════════════════════════════════════════
# 3. SbiPdfParser — TEXT_LAYER parse_text_content
# ═══════════════════════════════════════════════════════════════════════════════

_TEXT_SINGLE = """\
State Bank of India
Account Statement
01 Apr 2025  01 Apr 2025  BY TRANSFER-NEFT*ICIC0SF0002  TRANSFER FROM  5,000.00  27,527.17
"""

_TEXT_MULTI = """\
State Bank of India Account Statement
01 Apr 2025  01 Apr 2025  NEFT/TRANSFER IN                 REF001             10,000.00  30,000.00
10 Apr 2025  10 Apr 2025  UPI/DR/SWIGGY/FOOD               UPI001  450.00               29,550.00
15 Apr 2025  15 Apr 2025  ATM CASH WD BANGALORE            ATM001  2,000.00             27,550.00
20 Apr 2025  20 Apr 2025  IMPS/RECHARGE                    IMPS01  199.00               27,351.00
"""

class TestTextLayerParsing:
    def test_single_credit(self, parser, batch_id):
        result = parser.parse_text_content(batch_id, _TEXT_SINGLE)
        assert len(result.rows) >= 1

    def test_multi_txn_count(self, parser, batch_id):
        result = parser.parse_text_content(batch_id, _TEXT_MULTI)
        assert len(result.rows) == 4

    def test_upi_type(self, parser, batch_id):
        result = parser.parse_text_content(batch_id, _TEXT_MULTI)
        upi = [r for r in result.rows if r.txn_type_hint == TxnTypeHint.UPI]
        assert len(upi) >= 1

    def test_neft_type(self, parser, batch_id):
        result = parser.parse_text_content(batch_id, _TEXT_MULTI)
        neft = [r for r in result.rows if r.txn_type_hint == TxnTypeHint.NEFT]
        assert len(neft) >= 1

    def test_atm_type(self, parser, batch_id):
        result = parser.parse_text_content(batch_id, _TEXT_MULTI)
        atm = [r for r in result.rows if r.txn_type_hint == TxnTypeHint.ATM_WITHDRAWAL]
        assert len(atm) >= 1

    def test_imps_type(self, parser, batch_id):
        result = parser.parse_text_content(batch_id, _TEXT_MULTI)
        imps = [r for r in result.rows if r.txn_type_hint == TxnTypeHint.IMPS]
        assert len(imps) >= 1

    def test_empty_text_zero_rows(self, parser, batch_id):
        result = parser.parse_text_content(batch_id, "State Bank of India\n\nNo transactions.")
        assert len(result.rows) == 0

    def test_source_type_sbi_bank(self, parser, batch_id):
        result = parser.parse_text_content(batch_id, _TEXT_MULTI)
        assert all(r.source_type == SourceType.SBI_BANK for r in result.rows)

    def test_method_tagged(self, parser, batch_id):
        result = parser.parse_text_content(batch_id, _TEXT_SINGLE)
        assert result.method == ExtractionMethod.TEXT_LAYER


# ═══════════════════════════════════════════════════════════════════════════════
# 4. Helper functions (_normalise_cell, _infer_txn_type)
# ═══════════════════════════════════════════════════════════════════════════════

class TestHelpers:
    def test_normalise_none(self):
        assert _normalise_cell(None) == ""

    def test_normalise_empty(self):
        assert _normalise_cell("") == ""

    def test_normalise_newlines(self):
        assert _normalise_cell("13 May\n2025") == "13 May 2025"

    def test_normalise_multiple_whitespace(self):
        assert _normalise_cell("BY TRANSFER-\n  NEFT*000") == "BY TRANSFER- NEFT*000"

    def test_normalise_tabs(self):
        # Embedded tabs normalised to single space
        assert _normalise_cell("foo\tbar") == "foo bar"

    def test_infer_upi(self):
        assert _infer_txn_type("UPI/DR/1234") == TxnTypeHint.UPI

    def test_infer_neft(self):
        assert _infer_txn_type("NEFT CR TRANSFER") == TxnTypeHint.NEFT

    def test_infer_imps(self):
        assert _infer_txn_type("IMPS/001") == TxnTypeHint.IMPS

    def test_infer_atm(self):
        assert _infer_txn_type("ATM CASH WD SBI") == TxnTypeHint.ATM_WITHDRAWAL

    def test_infer_unknown(self):
        assert _infer_txn_type("OTHER TRANSFER") == TxnTypeHint.UNKNOWN

    def test_infer_case_insensitive_upi(self):
        assert _infer_txn_type("upi payment") == TxnTypeHint.UPI

    def test_infer_phonepay(self):
        assert _infer_txn_type("PHONEPE TRANSFER") == TxnTypeHint.UPI

    def test_infer_gpay(self):
        assert _infer_txn_type("GPAY PAYMENT") == TxnTypeHint.UPI


# ═══════════════════════════════════════════════════════════════════════════════
# 5. GenericCsvParser — old SBI pseudo-XLS (_read_sbi_pseudo_xls)
# ═══════════════════════════════════════════════════════════════════════════════

def _make_sbi_pseudo_xls(rows: list[dict]) -> bytes:
    """Build synthetic SBI pseudo-XLS bytes (tab-sep with 'quoted' cells)."""
    lines = [
        "Account Name       :\tMr. TEST USER",
        "Account Number     :\t_00000039146211724",
        "IFS Code           :\tSBIN0011348",
        "Balance on 1 Apr 2023        :\t5,000.00",
        "Start Date          :\t1 Apr 2023",
        "End Date            :\t31 Mar 2024",
        "'Txn Date'\t'Value Date'\t'Description'\t'Ref No./Cheque No.'\t'Debit'\t'Credit'\t'Balance'\t",
    ]
    for r in rows:
        line = (
            f"'{r['date']}'\t'{r['date']}'\t'{r['narr']}'\t'{r.get('ref', '')}'\t"
            f"'{r.get('debit', '')}'\t'{r.get('credit', '')}'\t'{r['balance']}'\t"
        )
        lines.append(line)
    return "\n".join(lines).encode("utf-8")


_PSEUDO_XLS_ROWS = [
    {"date": "2 Apr 2023", "narr": "BY TRANSFER-NEFT*ICIC0SF0002*DANTURTI", "ref": "TRANSFER FROM 4697249044303", "credit": "5,000.00", "balance": "12,232.55"},
    {"date": "19 Apr 2023", "narr": "TO TRANSFER-UPI/DR/310913519423/JADESWAM", "ref": "TRANSFER TO 4897692162094", "debit": "30.00", "balance": "12,202.55"},
    {"date": "20 Apr 2023", "narr": "TO TRANSFER-IMPS/DR/347641207339/ADYAR AN", "ref": "TRANSFER TO 4897693162093", "debit": "117.60", "balance": "12,084.95"},
]


class TestSbiPseudoXls:
    def test_read_returns_dataframe(self):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        df = GenericCsvParser._read_sbi_pseudo_xls(raw)
        assert df is not None
        assert len(df) == 3

    def test_columns_stripped_of_quotes(self):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        df = GenericCsvParser._read_sbi_pseudo_xls(raw)
        assert "Txn Date" in df.columns
        assert "'Txn Date'" not in df.columns

    def test_date_values_stripped_of_quotes(self):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        df = GenericCsvParser._read_sbi_pseudo_xls(raw)
        assert df["Txn Date"].iloc[0] == "2 Apr 2023"

    def test_balance_values_correct(self):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        df = GenericCsvParser._read_sbi_pseudo_xls(raw)
        assert df["Balance"].iloc[0] == "12,232.55"

    def test_credit_row(self):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        df = GenericCsvParser._read_sbi_pseudo_xls(raw)
        assert df["Credit"].iloc[0] == "5,000.00"

    def test_debit_row(self):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        df = GenericCsvParser._read_sbi_pseudo_xls(raw)
        assert df["Debit"].iloc[1] == "30.00"

    def test_returns_none_for_non_sbi(self):
        # A file without "Txn Date" header
        raw = b"Date\tDescription\tAmount\n01/01/2023\tTest\t100.00\n"
        result = GenericCsvParser._read_sbi_pseudo_xls(raw)
        assert result is None

    def test_preamble_rows_excluded(self):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        df = GenericCsvParser._read_sbi_pseudo_xls(raw)
        # Should not contain metadata rows like "Account Name"
        assert not any("Account Name" in str(v) for v in df["Txn Date"].values)


class TestSbiPseudoXlsFullParse:
    """Integration: _read_file → parse_dataframe via GenericCsvParser.extract()"""

    def test_full_extract_row_count(self, csv_parser, batch_id):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xls")
        assert len(result.rows) == 3

    def test_full_extract_source_type(self, csv_parser, batch_id):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xls")
        assert all(r.source_type == SourceType.SBI_BANK_CSV for r in result.rows)

    def test_full_extract_credit_amount(self, csv_parser, batch_id):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xls")
        credit_rows = [r for r in result.rows if r.raw_credit is not None]
        assert len(credit_rows) == 1
        assert credit_rows[0].raw_credit == "5,000.00"

    def test_full_extract_debit_amount(self, csv_parser, batch_id):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xls")
        debit_rows = [r for r in result.rows if r.raw_debit is not None]
        assert len(debit_rows) == 2

    def test_full_extract_balance(self, csv_parser, batch_id):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xls")
        assert result.rows[0].raw_balance == "12,232.55"

    def test_full_extract_txn_type_neft(self, csv_parser, batch_id):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xls")
        assert result.rows[0].txn_type_hint == TxnTypeHint.NEFT

    def test_full_extract_txn_type_upi(self, csv_parser, batch_id):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xls")
        assert result.rows[1].txn_type_hint == TxnTypeHint.UPI

    def test_full_extract_txn_type_imps(self, csv_parser, batch_id):
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xls")
        assert result.rows[2].txn_type_hint == TxnTypeHint.IMPS


# ═══════════════════════════════════════════════════════════════════════════════
# 6. GenericCsvParser — SBI XLSX (after OLE2 decryption → regular XLSX)
#    We test the column-mapping detection and dataframe parsing directly
# ═══════════════════════════════════════════════════════════════════════════════

def _make_sbi_xlsx_bytes(rows: list[dict]) -> bytes:
    """Build an in-memory XLSX with SBI column headers and optional preamble."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active

    # Preamble rows (simulates account info before transaction table)
    ws.append(["Account Name :", "Mr. TEST USER"])
    ws.append(["Account Number :", "00000039146211724"])
    ws.append(["IFS Code :", "SBIN0011348"])
    ws.append(["Nomination Registered :", "Yes"])
    ws.append(["Balance on 1 Apr 2025 :", "5,000.00"])
    ws.append([])  # blank line

    # Transaction table header
    ws.append(["Txn Date", "Value Date", "Description", "Ref No./Cheque No.", "Debit", "Credit", "Balance"])

    # Data rows
    for r in rows:
        ws.append([
            r["date"], r["date"], r["narr"], r.get("ref", ""),
            r.get("debit", ""), r.get("credit", ""), r["balance"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XLSX_ROWS = [
    {"date": "1 Apr 2025", "narr": "BY TRANSFER-NEFT*ICIC0SF0002*DANTURTI", "ref": "99509044300", "credit": "5,000.00", "balance": "27,527.17"},
    {"date": "3 Apr 2025", "narr": "TO TRANSFER-UPI/DR/509399466100/MACHA", "ref": "4897694162092", "debit": "5,800.00", "balance": "21,729.17"},
    {"date": "9 Apr 2025", "narr": "TO TRANSFER-IMPS/100327963734/NETFLI", "ref": "4897693162093", "debit": "199.00", "balance": "21,530.17"},
]


class TestSbiXlsxParsing:
    def test_xlsx_extract_row_count(self, csv_parser, batch_id):
        raw = _make_sbi_xlsx_bytes(_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="AccountStatement_24032026_094520.xlsx")
        assert len(result.rows) == 3

    def test_xlsx_extract_source_type(self, csv_parser, batch_id):
        raw = _make_sbi_xlsx_bytes(_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xlsx")
        assert all(r.source_type == SourceType.SBI_BANK_CSV for r in result.rows)

    def test_xlsx_credit_row(self, csv_parser, batch_id):
        raw = _make_sbi_xlsx_bytes(_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xlsx")
        assert result.rows[0].raw_credit == "5,000.00"
        assert result.rows[0].raw_debit is None

    def test_xlsx_debit_row(self, csv_parser, batch_id):
        raw = _make_sbi_xlsx_bytes(_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xlsx")
        assert result.rows[1].raw_debit == "5,800.00"
        assert result.rows[1].raw_credit is None

    def test_xlsx_balance(self, csv_parser, batch_id):
        raw = _make_sbi_xlsx_bytes(_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xlsx")
        assert result.rows[0].raw_balance == "27,527.17"

    def test_xlsx_date_preserved(self, csv_parser, batch_id):
        raw = _make_sbi_xlsx_bytes(_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi.xlsx")
        assert result.rows[0].raw_date == "1 Apr 2025"


# ═══════════════════════════════════════════════════════════════════════════════
# 7. Detector — filename + content-based detection for all SBI formats
# ═══════════════════════════════════════════════════════════════════════════════

class TestSbiDetector:
    def test_new_format_pdf_filename(self, detector):
        """AccountStatement_DDMMYYYY_HHMMSS.pdf → SBI_BANK (PDF format)."""
        result = detector.detect(
            filename="AccountStatement_24032026_094359.pdf",
            file_bytes=b"%PDF-1.4 encrypted",
        )
        assert result.source_type == SourceType.SBI_BANK
        assert result.confidence >= 0.85
        assert result.file_format == FileFormat.PDF

    def test_new_format_xlsx_filename_decrypted(self, detector):
        """AccountStatement_*.xlsx (decrypted → XLSX bytes) → SBI_BANK_CSV."""
        xlsx_bytes = b"PK\x03\x04" + b"\x00" * 100
        result = detector.detect(
            filename="AccountStatement_24032026_094520.xlsx",
            file_bytes=xlsx_bytes,
        )
        assert result.source_type == SourceType.SBI_BANK_CSV
        assert result.file_format == FileFormat.XLSX

    def test_sbi_old_pdf_content_detection(self, detector):
        """Old SBI PDFs contain 'State Bank of India' + 'Debit' + 'Credit'."""
        fake_pdf = b"%PDF-1.4 State Bank of India Account Txn Date Debit Credit Balance"
        result = detector.detect(filename="stmt.pdf", file_bytes=fake_pdf)
        assert result.source_type == SourceType.SBI_BANK

    def test_sbi_old_xls_header_scan(self, detector):
        """Old SBI pseudo-XLS: header scan via text fallback."""
        raw = _make_sbi_pseudo_xls(_PSEUDO_XLS_ROWS)
        result = detector.detect(filename="sbi_stmt.xls", file_bytes=raw)
        assert result.source_type == SourceType.SBI_BANK_CSV
        assert result.method == "header_scan"

    def test_sbi_bank_csv_signature(self, detector):
        """CSV with Txn Date / Description / Debit / Credit / Balance headers."""
        csv_text = b"Txn Date,Value Date,Description,Ref No.,Debit,Credit,Balance\n"
        result = detector.detect(filename="sbi.csv", file_bytes=csv_text)
        assert result.source_type == SourceType.SBI_BANK_CSV

    def test_sbionline_filename(self, detector):
        result = detector.detect(filename="sbionline_statement.xls", file_bytes=b"")
        assert result.source_type == SourceType.SBI_BANK_CSV

    def test_sbi_acc_filename(self, detector):
        result = detector.detect(filename="sbi_account_2025.pdf", file_bytes=b"%PDF-1.4")
        assert result.source_type == SourceType.SBI_BANK

    def test_new_format_various_timestamps(self, detector):
        """Different timestamps in the AccountStatement filename still match."""
        for fname in [
            "AccountStatement_01012025_120000.pdf",
            "AccountStatement_31122024_235959.xlsx",
        ]:
            ext_bytes = b"%PDF-1.4" if fname.endswith(".pdf") else b"PK\x03\x04" + b"\x00" * 10
            result = detector.detect(filename=fname, file_bytes=ext_bytes)
            assert result.source_type in (SourceType.SBI_BANK, SourceType.SBI_BANK_CSV), \
                f"Failed for {fname}: got {result.source_type}"


# ═══════════════════════════════════════════════════════════════════════════════
# 8. SbiPdfParser — new-format PDF positional table (empty headers except Balance)
# ═══════════════════════════════════════════════════════════════════════════════

# New-format pdfplumber output: header=['', '', '', '', '', '', 'Balance'],
# data cells use '-' for missing debit or credit.
_NEW_FMT_HEADER = ["", "", "", "", "", "", "Balance"]

def _new_fmt_table(data_rows: list[list[str]]) -> list[list[str]]:
    return [_NEW_FMT_HEADER] + data_rows


_NEW_FMT_CREDIT = _new_fmt_table([
    ["01/04/2025", "01/04/2025", "DEP TFR\nNEFT*ICIC0SF0002\nDANTURTI", "-", "-", "5,000.00", "27,527.17"],
])

_NEW_FMT_DEBIT = _new_fmt_table([
    ["03/04/2025", "03/04/2025", "WDL TFR\nUPI/DR/509399466100/MACHA", "-", "5,800.00", "-", "21,729.17"],
])

_NEW_FMT_MULTI = _new_fmt_table([
    ["01/04/2025", "01/04/2025", "DEP TFR\nNEFT*ICIC0SF0002\nDANTURTI", "-", "-", "5,000.00", "27,527.17"],
    ["03/04/2025", "03/04/2025", "WDL TFR\nUPI/DR/509399466100/MACHA",  "-", "5,800.00", "-", "21,729.17"],
    ["09/04/2025", "09/04/2025", "WDL TFR\nIMPS/100327963734/NETFLI",   "-", "199.00",   "-", "21,530.17"],
])


class TestNewFormatPdfPositional:
    """New SBI PDF format: positional column fallback + dash-as-null handling."""

    def test_credit_row_parsed(self, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_NEW_FMT_CREDIT])
        assert len(rows) == 1
        r = rows[0]
        assert r.raw_credit == "5,000.00"
        assert r.raw_debit is None

    def test_debit_row_parsed(self, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_NEW_FMT_DEBIT])
        assert len(rows) == 1
        r = rows[0]
        assert r.raw_debit == "5,800.00"
        assert r.raw_credit is None

    def test_dash_not_stored_as_amount(self, batch_id):
        """'-' placeholder must never appear as a raw_debit or raw_credit value."""
        rows = SbiPdfParser._rows_from_tables(batch_id, [_NEW_FMT_MULTI])
        for r in rows:
            assert r.raw_debit != "-"
            assert r.raw_credit != "-"

    def test_balance_extracted(self, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_NEW_FMT_CREDIT])
        assert rows[0].raw_balance == "27,527.17"

    def test_date_dd_mm_yyyy(self, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_NEW_FMT_CREDIT])
        assert rows[0].raw_date == "01/04/2025"

    def test_narration_newlines_stripped(self, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_NEW_FMT_CREDIT])
        assert "\n" not in rows[0].raw_narration

    def test_multi_row_count(self, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_NEW_FMT_MULTI])
        assert len(rows) == 3

    def test_sequential_row_numbers(self, batch_id):
        rows = SbiPdfParser._rows_from_tables(batch_id, [_NEW_FMT_MULTI])
        for i, r in enumerate(rows, start=1):
            assert r.row_number == i

    def test_page1_summary_table_skipped(self, batch_id):
        """Page 1 'Relationship Summary' table has no balance-only header — must be skipped."""
        summary_table = [
            ["Account Summary", "Details"],
            ["Account No.", "00000039146211724"],
            ["Branch", "MYSURU"],
        ]
        rows = SbiPdfParser._rows_from_tables(batch_id, [summary_table, _NEW_FMT_CREDIT])
        # Only the credit row from page 2 should be present
        assert len(rows) == 1
        assert rows[0].raw_credit == "5,000.00"

    def test_mixed_old_and_new_format_tables(self, batch_id):
        """Parser handles a mix of old-style named headers and new positional tables."""
        old_table = _make_table([
            ["1 Apr 2025", "1 Apr 2025", "OLD FORMAT NEFT", "REF001", "", "1,000.00", "10,000.00"],
        ])
        rows = SbiPdfParser._rows_from_tables(batch_id, [old_table, _NEW_FMT_DEBIT])
        assert len(rows) == 2


# ═══════════════════════════════════════════════════════════════════════════════
# 9. GenericCsvParser — new-format SBI XLSX (Date/Details headers, numeric amounts,
#    variable-length preamble that may exceed 40 rows)
# ═══════════════════════════════════════════════════════════════════════════════

def _make_new_sbi_xlsx(
    rows: list[dict],
    preamble_rows: int = 16,
) -> bytes:
    """Build an in-memory SBI new-format XLSX.

    The preamble length is configurable so we can test that even very long
    preambles (nominee details, account info, etc.) are handled correctly.
    Amounts are stored as *numbers* (no commas) matching real SBI new-format XLSX.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active

    # Variable-length preamble metadata rows
    labels = [
        "Account Holder Name  :  TEST USER",
        "Account Number  :  00000039146211724",
        "Address  :  123 Test Street, Mysuru",
        "IFS Code  :  SBIN0011348",
        "Branch  :  MYSURU MAIN BRANCH",
        "Account Type  :  Savings Account",
        "Clear Balance  :  36,756.18CR",
        "",
        "Monthly Avg Balance  :  0.00",
        "Interest Rate  :  2.50 % p.a.",
        "Product  :  Savings Account",
        "MICR Code  :  570002005",
        "Account Open Date  :  15/02/2020",
        "Nominee Name  :  XXXXXXXXXXX",
        "Nominee Date of Birth  :  XX/XX/XXXX",
        "Statement From  :  01-04-2025  to  24-03-2026",
    ]
    for i in range(preamble_rows):
        ws.append([labels[i % len(labels)]])

    # Transaction header — exact new-format column names
    ws.append(["Date", "Details", "Ref No/Cheque No", "Debit", "Credit", "Balance"])

    # Data rows — amounts as floats (no commas), matching real XLSX behaviour
    for r in rows:
        ws.append([
            r["date"],
            r["narr"],
            r.get("ref", ""),
            r.get("debit"),   # None when absent (credit row)
            r.get("credit"),  # None when absent (debit row)
            r["balance"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_NEW_XLSX_ROWS = [
    {"date": "01/04/2025", "narr": "DEP TFR\nNEFT*ICIC0SF0002\nDANTURTI", "ref": "",      "credit": 5000.00,  "balance": 27527.17},
    {"date": "03/04/2025", "narr": "WDL TFR\nUPI/DR/509399466100/MACHA",   "ref": "",      "debit":  5800.00,  "balance": 21729.17},
    {"date": "09/04/2025", "narr": "WDL TFR\nIMPS/100327963734/NETFLI",    "ref": "",      "debit":  199.00,   "balance": 21530.17},
]


class TestNewFormatSbiXlsx:
    """New-format SBI XLSX: Date/Details/Ref No/Cheque No headers, numeric amounts."""

    def test_row_count(self, csv_parser, batch_id):
        raw = _make_new_sbi_xlsx(_NEW_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="AccountStatement_24032026_094520.xlsx")
        assert len(result.rows) == 3

    def test_credit_amount_extracted(self, csv_parser, batch_id):
        raw = _make_new_sbi_xlsx(_NEW_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi_new.xlsx")
        assert result.rows[0].raw_credit is not None
        assert result.rows[0].raw_debit is None

    def test_debit_amount_extracted(self, csv_parser, batch_id):
        raw = _make_new_sbi_xlsx(_NEW_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi_new.xlsx")
        assert result.rows[1].raw_debit is not None
        assert result.rows[1].raw_credit is None

    def test_source_type_sbi_bank_csv(self, csv_parser, batch_id):
        raw = _make_new_sbi_xlsx(_NEW_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi_new.xlsx")
        assert all(r.source_type == SourceType.SBI_BANK_CSV for r in result.rows)

    def test_date_preserved(self, csv_parser, batch_id):
        raw = _make_new_sbi_xlsx(_NEW_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi_new.xlsx")
        assert result.rows[0].raw_date == "01/04/2025"

    def test_narration_newlines_stripped(self, csv_parser, batch_id):
        raw = _make_new_sbi_xlsx(_NEW_XLSX_ROWS)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi_new.xlsx")
        assert "\n" not in result.rows[0].raw_narration

    def test_long_preamble_50_rows(self, csv_parser, batch_id):
        """Preamble with 50 rows (beyond old 40-row limit) must still be detected."""
        raw = _make_new_sbi_xlsx(_NEW_XLSX_ROWS, preamble_rows=50)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi_new.xlsx")
        assert len(result.rows) == 3

    def test_long_preamble_80_rows(self, csv_parser, batch_id):
        """Preamble with 80 rows must still be detected."""
        raw = _make_new_sbi_xlsx(_NEW_XLSX_ROWS, preamble_rows=80)
        result = csv_parser.extract(batch_id, raw, ExtractionMethod.TABLE_EXTRACTION, filename="sbi_new.xlsx")
        assert len(result.rows) == 3

    def test_column_mapping_detected_as_sbi_xlsx_v1(self, batch_id):
        """detect_column_mapping maps Date/Details/Debit/Credit/Balance → sbi_xlsx_v1."""
        from modules.parser.parsers.generic_csv import detect_column_mapping
        headers = ["Date", "Details", "Ref No/Cheque No", "Debit", "Credit", "Balance"]
        mapping = detect_column_mapping(headers)
        assert mapping is not None
        assert mapping.format_fingerprint == "sbi_xlsx_v1"

    def test_column_mapping_date_format(self, batch_id):
        from modules.parser.parsers.generic_csv import detect_column_mapping
        headers = ["Date", "Details", "Ref No/Cheque No", "Debit", "Credit", "Balance"]
        mapping = detect_column_mapping(headers)
        assert mapping.date_format == "%d/%m/%Y"

    def test_detector_new_xlsx_headers(self, detector, batch_id):
        """Detector recognises new-format XLSX headers via content scan."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Details", "Ref No/Cheque No", "Debit", "Credit", "Balance"])
        buf = io.BytesIO()
        wb.save(buf)
        result = detector.detect(filename="statement.xlsx", file_bytes=buf.getvalue())
        assert result.source_type == SourceType.SBI_BANK_CSV
